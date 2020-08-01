import urllib.request
import urllib.parse
import openpyxl
import datetime
import re
import sys
from imaplib import IMAP4_SSL
import email

def fileDownload(): #скачивание файла
	print('Beginning file download...')
	urlSmena = '' 
	urlBase = ''
	urllib.request.urlretrieve(urlSmena,'') #путь сохранения файла !!!с расширением!!!
	urllib.request.urlretrieve(urlBase,'') #путь сохранения файла !!!с расширением!!!

def findInSmena(carName): #обращение к файлу смены
	carName = carName.upper()
	print(carName)
	smena = openpyxl.load_workbook(filename = '/home/hakerman/taxi/smena.xlsx') #путь сохранения файла !!!с расширением!!!
	sheet = smena['Октябрь19'] 			#название листа с данными 
	smenaRow = None #это костыли их не трогать
	smenaCol = None

	for row in sheet.iter_rows(2,2): 		#находим номер машины
		for cell in row:
			if carName == str(cell.value)[1:]:
				smenaCol = cell.column
	now = datetime.datetime.now()

	if now.hour > 12:				#начинаем считать день с пересменки
		day = now.day
	else:
		day = now.day-1

	for col in sheet.iter_cols(1,1):		#находим в листе день
		for cell in col:
			if cell.value != None:
				if day == cell.value.day:
					smenaRow = cell.row
	if smenaRow != None and smenaCol != None:
		surname = re.sub(r'[0-9\.]+', '', sheet.cell(smenaRow,smenaCol).value) #убираем все числа
		print(surname)
		return surname
	else:
		print('No name in cell')
		return None

def findInBase(name):	#обращение к файлу базы
	base = openpyxl.load_workbook(filename = '/home/hakerman/taxi/base.xlsx')
	sheet = base['Лист1']
	baseRow = None
	for col in sheet.iter_cols(2,2):	#находим фамилию
		for cell in col:
			if cell.value != None:
				if name == cell.value.split()[0]:
					baseRow = cell.row
	if baseRow != None:
		print(sheet.cell(baseRow,3).value) #возвращаем телефон
	else:
		print('No driver found')

def get_body(msg): #рекурсивная распаковка с помощью магии
    if msg.is_multipart():
        return get_body(msg.get_payload(0))
    else:
        return msg.get_payload(None,True)


#Код начинает работать от сюда
fileDownload()				#скачивание файла
YA_HOST = "imap.mail.ru"		#доступ к почте
YA_PORT = 
YA_USER = ""
YA_PASSWORD = ""
box = IMAP4_SSL(host=YA_HOST, port=YA_PORT)   	
box.login(user=YA_USER, password=YA_PASSWORD)
status, msgs = box.select('INBOX')
assert status == 'OK'
typ, data = box.search(None, '(UNSEEN)')	#непрочтенные, если все прочитаны, то программа закончит работу
for num in data[0].split():
	typ, message_data = box.fetch(num, '(RFC822)')	#последовательно читаем сообщения
	mail = email.message_from_bytes(message_data[0][1])
	message = urllib.parse.unquote(urllib.parse.quote(get_body(mail).decode('utf8'))) #декодируем из url в юникод
	print("\n" + message)
	name = findInSmena(message.split()[0].replace('198','').replace('82','')) #используем методы к базам
	findInBase(name)
box.close()	#Выход из почты
box.logout()

